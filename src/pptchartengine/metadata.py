"""Chart metadata persistence.

Belongs to: **metadata** lifecycle per ADR-0007 §1.
Realises: ADR-0004 §2 (metadata 5-piece round-trip contract)
          + ADR-0007 §3 (layered metadata strategy).

Single entrypoint for chart-level semantic metadata. Today only **layer 1**
(embedded workbook hidden sheet) is implemented; the layered strategy goal
in ADR-0007 §3 is:

    1. embedded workbook hidden sheet   ← layer 1, this module (default)
    2. chart semantic anchor / invisible shape
       ← layer 2, lives in ``semantic_anchor.py`` for shape-composition
         families that have no chart container
    3. custom XML part                  ← layer 3, requires PRD + compat check
    4. shape alt text / name            ← layer 4, selector hint only

This module holds the **single source of truth** for ``METADATA_SHEET_NAME``
and ``METADATA_SCHEMA_VERSION``; previously these were duplicated in
``api.py`` and ``parser.py``, a real silent-drift hazard.

Public:

- :data:`METADATA_SHEET_NAME`
- :data:`METADATA_SCHEMA_VERSION`
- :class:`ChartMetadataV1`
- :func:`write_chart_metadata`

Backward-compat (kept for existing callers):

- :func:`_write_embedded_metadata` — original signature preserved;
  internally delegates to layer 1 backend.

ADR-0007 §2: this module avoids Pydantic; uses standard-library
``dataclasses`` only.
"""

from __future__ import annotations

import io
import json
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional

from ._log import debug_print as print


# ---------------------------------------------------------------------------
# Authoritative constants (DO NOT redefine in api.py / parser.py)
# ---------------------------------------------------------------------------

METADATA_SHEET_NAME: str = "_pptchartengine_meta"
"""Name of the hidden sheet inside a chart's embedded workbook that holds
chart-level semantic metadata. Authoritative source; ``api.py`` and
``parser.py`` re-import this rather than redefining."""

METADATA_SCHEMA_VERSION: str = "2"
"""Current schema version written to ``B1`` of the metadata sheet.
Future schema changes must bump this AND maintain a backward-compat reader
in ``parser.py`` (ADR-0004 backward-compat clause)."""


# ---------------------------------------------------------------------------
# Public schema
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class ChartMetadataV1:
    """Chart-level semantic metadata payload (ADR-0007 §3 layer 1, schema v2).

    All fields are technical-layer only per ADR-0007 §2 — no business slot,
    no ``user_id``, no prompt.
    """

    categories_col: str
    series_config: List[Dict[str, Any]]
    chart_family: Optional[str] = None
    extra: Dict[str, Any] = field(default_factory=dict)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def write_chart_metadata(chart, metadata: ChartMetadataV1) -> None:
    """Persist :class:`ChartMetadataV1` into the chart's embedded workbook.

    Routes through layer 1 (embedded workbook hidden sheet). Failure is
    currently swallowed with a stderr-like log (preserves prior
    ``_write_embedded_metadata`` semantics); a future schema version should
    promote this to a typed :class:`pptchartengine.replace` style result or
    raise a dedicated ``MetadataWriteError``.
    """
    _write_workbook_hidden_sheet(
        chart=chart,
        categories_col=metadata.categories_col,
        series_config=metadata.series_config,
        chart_family=metadata.chart_family,
        extra=metadata.extra,
    )


# ---------------------------------------------------------------------------
# Backward-compat (kept until callers migrate to write_chart_metadata)
# ---------------------------------------------------------------------------


def _write_embedded_metadata(
    chart,
    categories_col: str,
    series_config: List[Dict],
    metadata: Optional[Dict] = None,
) -> None:
    """Legacy signature preserved for ``api.py`` / ``scatter.py`` /
    ``semantic_family.py`` call sites. Routes to layer 1 backend.

    ``metadata`` here is a free-form dict that historically carried
    ``chart_family`` plus arbitrary extra fields. We extract ``chart_family``
    if present, treat the whole dict as ``extra``.
    """
    chart_family = None
    if metadata is not None:
        chart_family = metadata.get("chart_family")
    _write_workbook_hidden_sheet(
        chart=chart,
        categories_col=categories_col,
        series_config=series_config,
        chart_family=chart_family,
        extra=metadata,
    )


# ---------------------------------------------------------------------------
# Layer 1 backend — embedded workbook hidden sheet
# ---------------------------------------------------------------------------


def _write_workbook_hidden_sheet(
    chart,
    categories_col: str,
    series_config: List[Dict[str, Any]],
    chart_family: Optional[str] = None,
    extra: Optional[Dict[str, Any]] = None,
) -> None:
    """Layer 1 implementation: write metadata into a hidden sheet inside
    chart's embedded xlsx workbook. Preserves the exact byte layout of the
    legacy ``_write_embedded_metadata`` so round-trip parsers stay happy."""
    try:
        from openpyxl import load_workbook

        chart_part = chart.part
        xlsx_part = chart_part.chart_workbook.xlsx_part
        xlsx_stream = io.BytesIO(xlsx_part.blob)
        wb = load_workbook(xlsx_stream)

        if METADATA_SHEET_NAME in wb.sheetnames:
            del wb[METADATA_SHEET_NAME]

        ws = wb.create_sheet(METADATA_SHEET_NAME)
        ws.sheet_state = "hidden"
        ws["A1"] = "schema_version"
        ws["B1"] = METADATA_SCHEMA_VERSION
        ws["A2"] = "categories_col"
        ws["B2"] = categories_col
        ws["A3"] = "series_count"
        ws["B3"] = len(series_config)
        ws["A4"] = "chart_family"
        ws["B4"] = chart_family
        ws["A5"] = "chart_metadata_json"
        # Legacy parity: the json blob is the full ``extra`` dict (which in
        # practice already carries chart_family); we don't add a second copy.
        ws["B5"] = json.dumps(extra, ensure_ascii=False) if extra else None
        ws.append([])
        ws.append([
            "series_index", "key", "name", "type",
            "axis", "grouping", "x_key", "size_key",
        ])

        for index, series in enumerate(series_config):
            ws.append([
                index,
                series.get("key"),
                series.get("name"),
                series.get("type"),
                series.get("axis"),
                series.get("grouping"),
                series.get("x_key"),
                series.get("size_key"),
            ])

        output_stream = io.BytesIO()
        wb.save(output_stream)
        xlsx_part._blob = output_stream.getvalue()

    except Exception as e:
        print(f"  ⚠️ 写入图表元数据失败: {e}")
