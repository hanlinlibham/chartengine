"""
公共 API - 简洁的高层接口

这是用户（您的工作室同事）唯一需要导入的模块。
隐藏所有实现细节。
"""

import json
from typing import List, Dict, Optional
from pptx.slide import Slide
from pptx.enum.chart import XL_CHART_TYPE
from pptx.chart.data import BubbleChartData, CategoryChartData, XyChartData
from pptx.util import Inches
import pandas as pd

from .builder import ChartBuilder
from .date_axis import format_category_label
from .metadata import (
    METADATA_SCHEMA_VERSION,
    METADATA_SHEET_NAME,
    _write_embedded_metadata,
)

XY_CHART_TYPES = ("scatter", "bubble")

# 导入样式模块
try:
    from .styles import StyleConfig, DEFAULT_STYLE_CONFIG
except ImportError:
    StyleConfig = None
    DEFAULT_STYLE_CONFIG = None

# 导入布局模块
try:
    from .layout import ChartLayoutConfig
except ImportError:
    ChartLayoutConfig = None


def create_combo_chart(
    slide: Slide,
    df: pd.DataFrame,
    categories_col: str,
    series_config: List[Dict],
    position: tuple = (Inches(1), Inches(2)),
    size: tuple = (Inches(8), Inches(4.5)),
    style_config=None,
    layout_config=None,
    metadata: Optional[Dict] = None,
):
    """
    创建组合图（支持 P(n,2) 任意组合）
    
    Args:
        slide: 幻灯片对象
        df: 数据 DataFrame
        categories_col: 分类列名（X 轴）
        series_config: 系列配置列表
            [
                {"key": "销售额", "name": "销售额", "type": "bar", "axis": "primary"},
                {"key": "增长率", "name": "增长率", "type": "line", "axis": "secondary"},
                {"key": "市场份额", "name": "市场份额", "type": "line", "axis": "secondary"},
            ]
        position: 图表位置 (left, top)
        size: 图表大小 (width, height)
        style_config: 样式配置对象（可选，默认使用 DEFAULT_STYLE_CONFIG）
                     可以是 StyleConfig 实例，或 None 使用默认样式
        layout_config: 布局配置对象（可选）
                      可以是 ChartLayoutConfig 实例，包含图例、轴配置
        
    Returns:
        Chart 对象
        
    Examples:
        >>> # 示例 1: 使用默认样式和布局
        >>> chart = create_combo_chart(
        ...     slide=slide,
        ...     df=df,
        ...     categories_col="日期",
        ...     series_config=[
        ...         {"key": "销售额", "name": "销售额", "type": "bar", "axis": "primary"},
        ...         {"key": "增长率", "name": "增长率", "type": "line", "axis": "secondary"},
        ...     ]
        ... )
        
        >>> # 示例 2: 自定义样式 + 布局
        >>> from pptchartengine import StyleConfig
        >>> from pptchartengine import (
        ...     ChartLayoutConfig,
        ...     LegendConfig,
        ...     CategoryAxisConfig,
        ... )
        >>> 
        >>> # 样式配置
        >>> custom_style = StyleConfig(
        ...     color_scheme="dark_only",
        ...     line_width_pt=1.5,
        ...     marker_style="none",
        ... )
        >>> 
        >>> # 布局配置
        >>> custom_layout = ChartLayoutConfig(
        ...     legend_config=LegendConfig(position="bottom", font_size_pt=10),
        ...     category_axis_config=CategoryAxisConfig(
        ...         is_date_axis=True,
        ...         major_unit_days=7,  # 每周显示一个刻度
        ...         number_format="yyyy-mm-dd",
        ...     ),
        ... )
        >>> 
        >>> chart = create_combo_chart(
        ...     slide=slide,
        ...     df=df,
        ...     categories_col="日期",
        ...     series_config=[...],
        ...     style_config=custom_style,
        ...     layout_config=custom_layout
        ... )
    
    Supported Combinations:
        - type: 'bar', 'column', 'line', 'area', 'scatter', 'bubble'
        - axis: 'primary', 'secondary'
        - 任意 (type1, axis1) + (type2, axis2) 的组合
        - 支持主轴多种类型，次轴多种类型

    Initial XY-family Support:
        - scatter / bubble 当前仅支持纯 XY 家族图表
        - 不支持与分类轴图表混搭，也不支持 scatter 与 bubble 互混
        - 所有 XY 系列必须共享同一组 X 数据并使用主轴
        
    Notes:
        - 分类（X轴）在 Excel 中占用 A 列
        - 系列数据从 B 列开始
        - 支持超过 25 个系列（AA, AB, ...）
        - 左轴标签在左侧，右轴标签在右侧，不会重叠
        - 默认样式：无标记点、1pt 线宽、深浅色交替
        - 默认布局：图例在底部、横轴普通分类轴
    """
    if not series_config:
        raise ValueError("series_config 不能为空")

    position = _normalize_position_tuple(position)
    size = _normalize_size_tuple(size)
    _validate_series_config(df, categories_col, series_config)
    
    # 2. 创建引导图表（写入全部系列数据到嵌入 Excel）
    chart = _bootstrap_chart(
        slide, df, categories_col, series_config, position, size
    )
    
    # ⭐ 核心修复：修正嵌入的 Excel 工作表中的日期数据
    # 如果分类列是日期类型，需要将 Excel 工作表中的文本日期转换为真实的日期数值
    _fix_embedded_excel_dates(chart, df, categories_col)
    _write_embedded_metadata(chart, categories_col, series_config, metadata=metadata)
    
    # 3. 使用构建器完成剩余工作（传递样式配置和布局配置）
    builder = ChartBuilder(
        chart, 
        df, 
        categories_col,
        style_config=style_config if style_config is not None else DEFAULT_STYLE_CONFIG,
        layout_config=layout_config
    )
    
    # 注意：引导图表已经创建了第一个系列，构建器会继续追加
    # 如果需要完全自定义，可以在 builder.clear_bootstrap_chart() 中清理
    
    return builder.build(series_config)


def _bootstrap_chart(
    slide: Slide,
    df: pd.DataFrame,
    categories_col: str,
    series_config: List[Dict],
    position: tuple,
    size: tuple,
):
    """
    创建引导图表（写入全部系列数据到嵌入 Excel）

    用途：
    - 激活 <c:plotArea>，使其可以通过 XML 访问
    - 创建初始的分类轴和值轴
    - 将所有系列数据写入嵌入 Excel（确保"编辑数据"不丢数据）
    - 图表 XML 结构后续由 ChartBuilder 重建

    Args:
        slide: 幻灯片对象
        df: 数据 DataFrame
        categories_col: 分类列名
        series_config: 全部系列配置列表
        position: (left, top)
        size: (width, height)

    Returns:
        Chart 对象
    """
    chart_family = _get_chart_family(series_config)
    if chart_family in XY_CHART_TYPES:
        chart_data, chart_type = _build_xy_chart_data(chart_family, df, categories_col, series_config)
    else:
        chart_data = CategoryChartData()

        # 设置分类（X轴）
        categories = df[categories_col].tolist()

        categories_bootstrap = [format_category_label(cat) for cat in categories]

        chart_data.categories = categories_bootstrap

        # 添加全部系列数据（确保嵌入 Excel 包含所有列）
        for series_cfg in series_config:
            chart_data.add_series(
                series_cfg["name"],
                df[series_cfg["key"]].tolist()
            )

        # 使用第一个系列的类型决定引导图表类型
        chart_type = _get_chart_type(series_config[0].get("type", "bar"))

    # 创建图表
    left, top = _normalize_position_tuple(position)
    width, height = _normalize_size_tuple(size)
    graphic_frame = slide.shapes.add_chart(
        chart_type, left, top, width, height, chart_data
    )

    return graphic_frame.chart


def _get_chart_type(type_str: str) -> XL_CHART_TYPE:
    """将图表类型字符串转换为 XL_CHART_TYPE 枚举"""
    type_map = {
        "bar": XL_CHART_TYPE.COLUMN_CLUSTERED,
        "column": XL_CHART_TYPE.COLUMN_CLUSTERED,
        "line": XL_CHART_TYPE.LINE,
        "area": XL_CHART_TYPE.AREA,
        "scatter": XL_CHART_TYPE.XY_SCATTER,
        "bubble": XL_CHART_TYPE.BUBBLE,
    }
    return type_map.get(type_str.lower(), XL_CHART_TYPE.COLUMN_CLUSTERED)


def _build_xy_chart_data(
    chart_family: str,
    df: pd.DataFrame,
    categories_col: str,
    series_config: List[Dict],
):
    chart_data = XyChartData() if chart_family == "scatter" else BubbleChartData()

    for series_cfg in series_config:
        series = chart_data.add_series(series_cfg["name"])
        x_values = pd.to_numeric(df[series_cfg.get("x_key", categories_col)], errors="raise").tolist()
        y_values = pd.to_numeric(df[series_cfg["key"]], errors="raise").tolist()

        if chart_family == "bubble":
            size_values = pd.to_numeric(df[series_cfg["size_key"]], errors="raise").tolist()
            for x_value, y_value, size_value in zip(x_values, y_values, size_values):
                series.add_data_point(x_value, y_value, size_value)
        else:
            for x_value, y_value in zip(x_values, y_values):
                series.add_data_point(x_value, y_value)

    return chart_data, _get_chart_type(chart_family)


def _normalize_position_tuple(position: tuple):
    if len(position) != 2:
        raise ValueError("position 必须是 (left, top)")
    return tuple(_normalize_measure(value) for value in position)


def _normalize_size_tuple(size: tuple):
    if len(size) != 2:
        raise ValueError("size 必须是 (width, height)")
    return tuple(_normalize_measure(value) for value in size)


def _normalize_measure(value):
    if hasattr(value, "emu"):
        return value
    try:
        numeric = float(value)
    except Exception:
        return value
    if numeric < 1000:
        return Inches(numeric)
    return int(round(numeric))


def _group_series(series_config: List[Dict]) -> Dict[tuple, List[Dict]]:
    """按 (type, axis) 分组系列"""
    from collections import defaultdict
    groups = defaultdict(list)
    for cfg in series_config:
        key = (cfg.get("type", "bar"), cfg.get("axis", "primary"))
        groups[key].append(cfg)
    return dict(groups)


def _get_chart_family(series_config: List[Dict]) -> str:
    normalized_types = {str(cfg.get("type", "bar")).lower() for cfg in series_config}
    xy_types = normalized_types.intersection(XY_CHART_TYPES)

    if not xy_types:
        return "category"
    if xy_types != normalized_types:
        return "mixed"
    if len(xy_types) > 1:
        return "mixed_xy"
    return next(iter(xy_types))


def _validate_series_config(df: pd.DataFrame, categories_col: str, series_config: List[Dict]) -> None:
    chart_family = _get_chart_family(series_config)
    if chart_family == "mixed":
        raise ValueError("初始 scatter/bubble 支持不允许与 bar/line/area 混搭")
    if chart_family == "mixed_xy":
        raise ValueError("初始 XY 支持要求 scatter 和 bubble 分开使用")
    if chart_family in XY_CHART_TYPES:
        _validate_xy_series_config(df, categories_col, series_config, chart_family)


def _validate_xy_series_config(
    df: pd.DataFrame,
    categories_col: str,
    series_config: List[Dict],
    chart_family: str,
) -> None:
    shared_x_key = None
    shared_x_values = None

    for series_cfg in series_config:
        if series_cfg.get("axis", "primary") != "primary":
            raise ValueError("初始 scatter/bubble 支持仅允许 primary 轴")

        x_key = series_cfg.get("x_key", categories_col)
        if x_key not in df.columns:
            raise ValueError(f"散点/气泡图缺少 X 列: {x_key}")
        if series_cfg["key"] not in df.columns:
            raise ValueError(f"散点/气泡图缺少 Y 列: {series_cfg['key']}")

        x_values = df[x_key].tolist()
        if shared_x_key is None:
            shared_x_key = x_key
            shared_x_values = x_values
        elif x_key != shared_x_key or x_values != shared_x_values:
            raise ValueError("初始 scatter/bubble 支持要求所有系列共享同一个 X 列和相同 X 数据")

        pd.to_numeric(df[x_key], errors="raise")
        pd.to_numeric(df[series_cfg["key"]], errors="raise")

        if chart_family == "bubble":
            size_key = series_cfg.get("size_key")
            if not size_key:
                raise ValueError("bubble 系列必须提供 size_key")
            if size_key not in df.columns:
                raise ValueError(f"气泡图缺少 size 列: {size_key}")
            pd.to_numeric(df[size_key], errors="raise")


def _fix_embedded_excel_dates(chart, df: pd.DataFrame, categories_col: str):
    """
    修正嵌入的 Excel 工作表中的日期数据
    
    新方案：将日期格式化为字符串标签（如 "2024/01"）
    这样 PowerPoint 就会正确显示，而不会出现 1900 年问题
    
    Args:
        chart: python-pptx Chart 对象
        df: 数据 DataFrame
        categories_col: 分类列名
    """
    # 检查是否为日期类型
    if not pd.api.types.is_datetime64_any_dtype(df[categories_col]):
        print(f"  → 分类列不是日期类型，跳过 Excel 工作表修正")
        return  # 不是日期类型，无需修正
    
    print(f"\n🔧 修正嵌入的 Excel 工作表日期数据（转换为格式化字符串）...")
    
    try:
        from datetime import datetime
        from openpyxl import load_workbook
        import io
        
        # 获取嵌入的 Excel 数据
        chart_part = chart.part
        xlsx_part = chart_part.chart_workbook.xlsx_part
        
        print(f"  → 找到嵌入的 Excel 工作表")
        
        # 将 Excel blob 加载为 openpyxl workbook
        xlsx_stream = io.BytesIO(xlsx_part.blob)
        wb = load_workbook(xlsx_stream)
        ws = wb.active
        
        print(f"  → 工作表行数: {ws.max_row}, 列数: {ws.max_column}")
        
        # 获取日期数据
        categories = df[categories_col].tolist()
        
        print(f"  → 准备修正 {len(categories)} 个日期值")
        print(f"  → 第一个值: {categories[0]} (类型: {type(categories[0])})")
        
        # 修正 A 列（分类列）的数据 - 转换为格式化字符串
        # Excel 工作表的第一行是表头，数据从第二行开始
        fixed_count = 0
        for i, cat_value in enumerate(categories, start=2):
            if hasattr(cat_value, 'to_pydatetime'):
                cat_value = cat_value.to_pydatetime()
            
            date_str = format_category_label(cat_value, "yyyy/mm")
            if date_str != str(cat_value):
                ws.cell(row=i, column=1).value = date_str
                fixed_count += 1
        
        print(f"  → 已修正 {fixed_count} 个单元格")
        print(f"  → 示例：{format_category_label(categories[0], 'yyyy/mm') if categories else 'N/A'}")
        
        # 将修改后的 workbook 写回 blob
        output_stream = io.BytesIO()
        wb.save(output_stream)
        xlsx_part._blob = output_stream.getvalue()
        
        print(f"  ✅ 嵌入 Excel 工作表修正完成（{fixed_count} 个日期值转换为格式化字符串）")
        
    except Exception as e:
        print(f"  ⚠️ 修正嵌入 Excel 工作表失败: {e}")
        import traceback
        traceback.print_exc()


# Metadata persistence moved to ``pptchartengine.metadata`` (2026-05-24).
# ``_write_embedded_metadata`` is now re-exported from this module via the
# top-of-file ``from .metadata import _write_embedded_metadata``, so callers
# that do ``from .api import _write_embedded_metadata`` continue working
# unchanged. The legacy 60-line implementation that used to live here has
# been moved verbatim to ``metadata._write_workbook_hidden_sheet`` and is
# now the single source of truth.


# ============================================================================
# 便捷函数：向后兼容
# ============================================================================

def create_dual_axis_chart(
    slide: Slide,
    df: pd.DataFrame,
    categories_col: str,
    bar_columns: List[str],
    bar_names: List[str],
    line_columns: List[str],
    line_names: List[str],
    position: tuple = (Inches(1), Inches(2)),
    size: tuple = (Inches(8), Inches(4.5)),
):
    """
    便捷函数：创建双轴组合图（柱状图 + 折线图）
    
    这是向后兼容的 API，与旧的 xml_chart_patcher 接口一致。
    
    Example:
        >>> create_dual_axis_chart(
        ...     slide=slide,
        ...     df=df,
        ...     categories_col="日期",
        ...     bar_columns=["销售额", "成本"],
        ...     bar_names=["销售额", "成本"],
        ...     line_columns=["利润率"],
        ...     line_names=["利润率"],
        ... )
    """
    # 构建统一的 series_config
    series_config = []
    
    # 主轴柱状图
    for col, name in zip(bar_columns, bar_names):
        series_config.append({
            "key": col,
            "name": name,
            "type": "bar",
            "axis": "primary"
        })
    
    # 次轴折线图
    for col, name in zip(line_columns, line_names):
        series_config.append({
            "key": col,
            "name": name,
            "type": "line",
            "axis": "secondary"
        })
    
    # 调用统一的 API
    return create_combo_chart(
        slide, df, categories_col, series_config, position, size
    )
